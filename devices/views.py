from pathlib import Path
import os
import json
import shutil
import subprocess

from datetime import datetime, date, timedelta

from django.conf import settings
from django.contrib import messages
from django.contrib.auth import (
    authenticate,
    login,
    logout,
    update_session_auth_hash,
)
from django.contrib.auth.decorators import login_required
from django.contrib.auth.forms import PasswordChangeForm
from django.contrib.auth.models import User
from django.contrib.admin.views.decorators import staff_member_required

from django.db.models import (
    Count,
    Q,
    Case,
    When,
    Value,
    IntegerField,
)

from django.http import (
    FileResponse,
    Http404,
    JsonResponse,
)

from django.shortcuts import (
    render,
    redirect,
    get_object_or_404,
)

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


from .models import (
    Device,
    Reparatur,
    PracticeSettings,
    DeviceDocument,
    GeneralDocument,
    UserProfile,
    HomeInformation,
    HomeImage,
    Standort,
    Filterwechsel,
    Geraetart,
    TechnicianDocument,
    DocumentType,
    DashboardSettings,
    CompanyInformation,
    AuditLog,
    ContactInformation,
    ContactImage,
    DashboardWidget,
)


from .forms import (
    DeviceForm,
    PracticeSettingsForm,
    DeviceDocumentForm,
    GeneralDocumentForm,
    FilterwechselForm,
    GeraetartForm,
    TechnicianDocumentForm,
    HomeInformationForm,
    ReparaturForm,
    ReparaturBearbeitenForm,
)


def login_view(request):


    if request.method == "POST":

        username = request.POST.get("username")
        password = request.POST.get("password")


        user = authenticate(
            request,
            username=username,
            password=password
        )


        if user:

            login(request, user)

            return redirect("home")


        else:

            return render(
                request,
                "login.html",
                {
                    "error": "Benutzername oder Passwort falsch"
                }
            )


    return render(
        request,
        "login.html"
    )
@login_required
def home(request):

    create_daily_backup(request)

    home, created = HomeInformation.objects.get_or_create(
        id=1
    )

    return render(
        request,
        "devices/home.html",
        {
            "home": home,
        },
    )
@login_required
def home_edit(request):

    if not request.user.is_superuser:
        return redirect("permission_denied")

    home, created = HomeInformation.objects.get_or_create(id=1)

    if request.method == "POST":

        form = HomeInformationForm(
            request.POST,
            instance=home
        )

        if form.is_valid():

            home = form.save()

            # حذف الصور المحددة
            if "delete_selected" in request.POST:

                delete_ids = request.POST.getlist("delete_images")

                HomeImage.objects.filter(
                    id__in=delete_ids,
                    home=home
                ).delete()

            # إضافة الصور الجديدة
            images = request.FILES.getlist("images")

            for image in images:

                HomeImage.objects.create(
                    home=home,
                    image=image
                )

            messages.success(
                request,
                "Startseite erfolgreich aktualisiert."
            )

            return redirect("home_edit")

    else:

        form = HomeInformationForm(
            instance=home
        )

    return render(
        request,
        "devices/home_edit.html",
        {
            "form": form,
            "home": home,
        },
    )
@login_required
def device_list(request):

    search = request.GET.get("search", "")
    practice = request.GET.get("practice", "")
    status = request.GET.get("status", "")
    sort = request.GET.get("sort", "")
    geraetart = request.GET.get("geraetart", "")

    devices = Device.objects.all()



    # =====================
    # Suche
    # =====================

    if search:

        devices = devices.filter(

            Q(name__icontains=search) |
            Q(inventory_number__icontains=search) |
            Q(serial_number__icontains=search) |
            Q(practice__name__icontains=search)

        )



    # =====================
    # Standort Filter
    # =====================

    if practice:

        devices = devices.filter(
            practice_id=practice
        )

    # =====================
    # Geräteart Filter
    # =====================

    if geraetart:

        devices = devices.filter(
            geraetart_id=geraetart
        )

    # =====================
    # Status Filter
    # =====================

    if status:

        devices = devices.filter(
            status=status
        )



    # =====================
    # Sortierung
    # =====================

    if sort == "operating_hours":

        devices = devices.order_by(
            "operating_hours"
        )


    elif sort == "-operating_hours":

        devices = devices.order_by(
            "-operating_hours"
        )



    elif sort == "next_stk":

        devices = devices.order_by(
            "next_stk"
        )



    elif sort == "-next_stk":

        devices = devices.order_by(
            "-next_stk"
        )



    elif sort == "practice":

        devices = devices.order_by(
            "practice__name"
        )



    elif sort == "-practice":

        devices = devices.order_by(
            "-practice__name"
        )



    else:

        devices = sorted(
            devices,
            key=lambda d:(
                d.practice.name if d.practice else "",
                d.geraetart.name if d.geraetart else "",
                d.inventory_number
            )
        )
    standorte = Standort.objects.filter(active=True).order_by("name")
    geraetarten= Geraetart.objects.filter(aktiv=True).order_by("name")
    return render(

        request,

        "devices/device_list.html",

        {

            "devices": devices,
            "search": search,
            "practice": practice,
            "status": status,
            "geraetart": geraetart,
            "standorte": standorte,
            "geraetarten": geraetarten,

        }

    )

@login_required
def device_detail(request, device_id):

    device = get_object_or_404(
        Device,
        id=device_id
    )

    document_types = DocumentType.objects.all().order_by("order")

    documents = DeviceDocument.objects.filter(
        device=device
    ).select_related("document_type").order_by(
        "document_type__name",
        "-upload_date"
    )

    return render(
        request,
        "devices/device_detail.html",
        {
            "device": device,
            "document_types": document_types,
            "documents": documents,
            "filterwechsel": device.filterwechsel.all().order_by("-datum"),
        }
    )


@login_required
def device_create(request):

    if request.method == "POST":

        form = DeviceForm(
            request.POST,
            request.FILES
        )


        if form.is_valid():

            device = form.save()


            AuditLog.objects.create(

                user=request.user,

                action="CREATE",

                model_name="Gerät",

                object_id=device.id,

                description=(
                    f"Gerät {device.name} "
                    f"({device.inventory_number}) "
                    "hinzugefügt."
                )

            )


            messages.success(
                request,
                "Gerät erfolgreich hinzugefügt."
            )


            return redirect(
                "device_list"
            )


    else:

        form = DeviceForm()


    return render(

        request,

        "devices/device_form.html",

        {
            "form": form
        }

    )

@login_required    
def device_create_geraetart(request, geraetart_id):

    geraetart = get_object_or_404(
        Geraetart,
        id=geraetart_id
    )

    if request.method == "POST":

        form = DeviceForm(
            request.POST,
            request.FILES,
            geraetart=geraetart
        )

        if form.is_valid():

            device = form.save(commit=False)

            device.geraetart = geraetart
            device.name = geraetart.name

            device.save()

            return redirect(
                "device_geraetart",
                geraetart_id=geraetart.id
            )

    else:

        form = DeviceForm(
            geraetart=geraetart
        )

    return render(
        request,
        "devices/device_form.html",
        {
            "form": form,
            "geraetart": geraetart,
        }
    )

@login_required        
def device_geraetart(request, geraetart):

    
    search = request.GET.get("search", "")
    practice = request.GET.get("practice", "")


    devices = Device.objects.filter(
            geraetart=geraetart
    )


    if search:
        devices = devices.filter(
            Q(name__icontains=search) |
            Q(inventory_number__icontains=search) |
            Q(serial_number__icontains=search)
        )


    if practice:
        devices = devices.filter(
            practice=practice
        )


    devices = devices.order_by("inventory_number")


    return render(
        request,
        "devices/device_category.html",
        {
            "devices": devices,
            "geraetart": geraetart,
            "search": search,
            "practice": practice,
        }
    )
@login_required
def contact(request):

    settings = PracticeSettings.objects.first()

    kontakt = ContactInformation.objects.first()


    return render(
        request,
        "devices/contact.html",
        {
            "settings": settings,
            "kontakt": kontakt,
        }
    )

@login_required
def faellige_stk(request):

    heute = date.today()
    grenze = heute + timedelta(days=30)

    devices = Device.objects.filter(
        next_stk__isnull=False,
        next_stk__lte=grenze
    ).order_by("next_stk")


    return render(
        request,
        "devices/faellige_stk.html",
        {
            "devices": devices,
            "heute": heute,
        }
    )

@login_required
def faellige_mtk(request):

    heute = date.today()
    grenze = heute + timedelta(days=30)

    devices = Device.objects.filter(
        next_mtk__isnull=False,
        next_mtk__lte=grenze
    ).order_by("next_mtk")


    return render(
        request,
        "devices/faellige_mtk.html",
        {
            "devices": devices,
            "heute": heute,
        }
    )


@login_required
def faellige_dguv(request):

    heute = date.today()
    grenze = heute + timedelta(days=30)

    devices = Device.objects.filter(
        next_dguv__isnull=False,
        next_dguv__lte=grenze
    ).order_by("next_dguv")


    return render(
        request,
        "devices/faellige_dguv.html",
        {
            "devices": devices,
            "heute": heute,
        }
    )

@login_required
def reparatur(request):

    reparaturen = Reparatur.objects.filter(
        status__in=["Offen", "In Bearbeitung"]
    ).order_by("-datum")

    return render(
        request,
        "devices/reparatur.html",
        {
            "reparaturen": reparaturen
        }
    )



@login_required
def search_reparatur_device(request):

    query = request.GET.get("q", "")

    devices = Device.objects.filter(
        Q(inventory_number__icontains=query) |
        Q(serial_number__icontains=query) |
        Q(name__icontains=query)
    ).order_by("inventory_number")[:10]

    data = []

    for device in devices:

        data.append({
            "id": device.id,
            "inventory": device.inventory_number,
            "name": device.name,
            "serial": device.serial_number,
            "standort": device.practice.name,
        })

    return JsonResponse(data, safe=False)

@login_required
def reparatur_create(request):

    if request.method == "POST":

        form = ReparaturForm(request.POST)


        if form.is_valid():


            reparatur = form.save(commit=False)


            # Gerät aus Formular übernehmen
            device = form.cleaned_data.get("geraet")


            reparatur.geraet = device


            # Standard Status
            reparatur.status = "Offen"


            reparatur.save()



            AuditLog.objects.create(

                user=request.user,

                action="CREATE",

                model_name="Reparatur",

                object_id=reparatur.id,

                description=(

                    f"Reparatur für Gerät "
                    f"{device.name} "
                    f"(Inventarnummer: {device.inventory_number}) "
                    "erstellt."

                )

            )



            messages.success(

                request,

                "Reparatur erfolgreich erstellt."

            )



            return redirect(
                "reparatur"
            )


    else:


        form = ReparaturForm()



    return render(

        request,

        "devices/reparatur_form.html",

        {
            "form": form,
        },

    )


@login_required
def reparatur_detail(request, pk):

    reparatur = get_object_or_404(
        Reparatur,
        pk=pk
    )


    readonly = request.GET.get("readonly") == "1"


    if readonly:

        form = ReparaturBearbeitenForm(
            instance=reparatur
        )

        return render(

            request,

            "devices/reparatur_detail.html",

            {
                "reparatur": reparatur,
                "form": form,
                "readonly": True,
            }

        )



    if request.method == "POST":


        form = ReparaturBearbeitenForm(

            request.POST,

            request.FILES,

            instance=reparatur

        )


        if form.is_valid():


            reparatur = form.save(
                commit=False
            )


            if request.POST.get(
                "reparaturbericht-clear"
            ):

                if reparatur.reparaturbericht:

                    reparatur.reparaturbericht.delete(
                        save=False
                    )

                reparatur.reparaturbericht = None



            if request.POST.get(
                "reparaturbild-clear"
            ):

                if reparatur.reparaturbild:

                    reparatur.reparaturbild.delete(
                        save=False
                    )

                reparatur.reparaturbild = None



            reparatur.status = request.POST.get(
                "status"
            )


            reparatur.save()



            AuditLog.objects.create(

                user=request.user,

                action="UPDATE",

                model_name="Reparatur",

                object_id=reparatur.id,

                description=(

                    f"Reparatur für Gerät "
                    f"{reparatur.geraet.name} "
                    f"(Inventarnummer: {reparatur.geraet.inventory_number}) "
                    "geändert."

                )

            )



            messages.success(

                request,

                "Reparatur erfolgreich gespeichert."

            )


            return redirect(
                "reparatur"
            )


    else:


        form = ReparaturBearbeitenForm(
            instance=reparatur
        )


    return render(

        request,

        "devices/reparatur_detail.html",

        {
            "reparatur": reparatur,
            "form": form,
            "readonly": False,
        }

    )
@login_required
def reparatur_uebersicht(request):

    reparaturen = Reparatur.objects.filter(
        status__in=["Offen", "In Bearbeitung"]
    )

    status = request.GET.get("status")

    if status:
        reparaturen = reparaturen.filter(status=status)

    reparaturen = reparaturen.order_by("-datum")

    return render(
        request,
        "devices/reparatur.html",
        {
            "reparaturen": reparaturen,
            "title": "Reparaturübersicht",
            "count": reparaturen.count(),
        },
    )
@login_required
def reparatur_historie(request):

    reparaturen = Reparatur.objects.filter(
        status="Erledigt"
    ).order_by("-datum")


    search = request.GET.get("search")


    if search:

        reparaturen = reparaturen.filter(

            Q(geraet__inventory_number__icontains=search) |

            Q(geraet__serial_number__icontains=search) |

            Q(beschreibung__icontains=search) |

            Q(ausfuehrung__icontains=search)

        )


    return render(
        request,
        "devices/reparatur_historie.html",
        {
            "reparaturen": reparaturen,
            "search": search
        }
    )
@login_required
def reparatur_historie_detail(request, pk):

    reparatur = get_object_or_404(
        Reparatur,
        pk=pk
    )

    form = ReparaturBearbeitenForm(instance=reparatur)

    return render(
        request,
        "devices/reparatur_detail.html",
        {
            "reparatur": reparatur,
            "form": form,
            "readonly": True,
        }
    )
@login_required
def repair_delete(request, repair_id):

    reparatur = get_object_or_404(
        Reparatur,
        id=repair_id
    )


    if request.method == "POST":


        AuditLog.objects.create(

            user=request.user,

            action="DELETE",

            model_name="Reparatur",

            object_id=reparatur.id,

            description=(

                f"Reparatur für Gerät "
                f"{reparatur.geraet.name} "
                f"(Inventarnummer: {reparatur.geraet.inventory_number}) "
                "gelöscht."

            )

        )


        reparatur.delete()


        messages.success(

            request,

            "Reparatur gelöscht."

        )


    return redirect(
        "reparatur"
    )

@login_required
def device_search(request):

    query = request.GET.get("q", "")

    devices = Device.objects.none()

    if query:
        devices = Device.objects.filter(
            Q(name__icontains=query) |
            Q(inventory_number__icontains=query) |
            Q(serial_number__icontains=query)
        ).order_by("name")

    return render(
        request,
        "devices/device_search.html",
        {
            "devices": devices,
            "query": query,
        }
    )
def settings_view(request):
    return render(
        request,
        "devices/settings.html"
    )
@login_required
def practice_settings_edit(request):

    if not request.user.is_superuser:
        return redirect("permission_denied")


    settings = PracticeSettings.objects.first()

    if request.method == "POST":

        form = PracticeSettingsForm(
            request.POST,
            instance=settings
        )

        if form.is_valid():
            form.save()

            return render(
                request,
                "devices/settings.html"
            )

    else:

        form = PracticeSettingsForm(
            instance=settings
        )


    return render(
        request,
        "devices/practice_settings_edit.html",
        {
            "form": form
        }
    )
@login_required
def dashboard(request):

    heute = date.today()
    grenze = heute + timedelta(days=30)

    devices = Device.objects.all()

    dashboard_settings, created = DashboardSettings.objects.get_or_create(id=1)

    # =====================
    # Suche
    # =====================

    search = request.GET.get("search")

    if search:
        devices = devices.filter(
            Q(name__icontains=search) |
            Q(inventory_number__icontains=search) |
            Q(serial_number__icontains=search) |
            Q(practice__name__icontains=search)
        )

    # =====================
    # Filter Status
    # =====================

    status = request.GET.get("status")

    if status:
        devices = devices.filter(status=status)

    # =====================
    # Sortierung
    # =====================

    sort = request.GET.get("sort")

    if sort == "inventory_number":
        devices = devices.order_by("inventory_number")

    elif sort == "-inventory_number":
        devices = devices.order_by("-inventory_number")

    elif sort == "operating_hours":
        devices = devices.order_by("operating_hours")

    elif sort == "-operating_hours":
        devices = devices.order_by("-operating_hours")

    elif sort == "practice":
        devices = devices.order_by("practice")

    elif sort == "-practice":
        devices = devices.order_by("-practice")

    elif sort == "next_stk":
        devices = devices.order_by("next_stk")

    elif sort == "-next_stk":
        devices = devices.order_by("-next_stk")

    # =====================
    # Standorte
    # =====================

    luebeck_total = Device.objects.filter(
        practice__name="Lübeck"
    ).count()

    ratzeburg_total = Device.objects.filter(
        practice__name="Ratzeburg"
    ).count()

    # =====================
    # Reparaturen
    # =====================

    offene_reparaturen = Reparatur.objects.filter(
        status="Offen"
    ).count()

    reparaturen_bearbeitung = Reparatur.objects.filter(
        status="In Bearbeitung"
    ).count()

    # =====================
    # Wartungen
    # =====================

    faellige_stk = Device.objects.filter(
        next_stk__isnull=False,
        next_stk__lte=grenze
    ).count()

    faellige_mtk = Device.objects.filter(
        next_mtk__isnull=False,
        next_mtk__lte=grenze
    ).count()

    faellige_dguv = Device.objects.filter(
        next_dguv__isnull=False,
        next_dguv__lte=grenze
    ).count()

    # =====================
    # Geräteübersicht
    # =====================

    geraete_uebersicht = Device.objects.values(
        "geraetart__name",
        "practice__name"
    ).annotate(
        anzahl=Count("id")
    ).order_by(
        "geraetart__name",
        "practice__name"
    )

    # =====================
    # Dashboard Widgets
    # =====================

    dashboard_widgets = DashboardWidget.objects.filter(
        visible=True
    ).order_by("order")

    for widget in dashboard_widgets:

        if widget.widget_type == "devices":

            qs = Device.objects.all()

            if widget.standort:
                qs = qs.filter(
                    practice=widget.standort
                )

            if widget.geraetart:
                qs = qs.filter(
                    geraetart=widget.geraetart
                )

            widget.value = qs.count()

        elif widget.widget_type == "devices_overview":
            widget.value = Device.objects.count()

        elif widget.widget_type == "repairs_open":
            widget.value = Reparatur.objects.filter(
                status="Offen"
            ).count()

        elif widget.widget_type == "repairs_progress":
            widget.value = Reparatur.objects.filter(
                status="In Bearbeitung"
            ).count()

        elif widget.widget_type == "repairs_done":

            widget.value = Reparatur.objects.filter(
                status="Erledigt"
            ).count()


        elif widget.widget_type == "repair_history":

            widget.value = Reparatur.objects.count()

        elif widget.widget_type == "filter_history":
            widget.value = Filterwechsel.objects.count()

        elif widget.widget_type == "stk":
            widget.value = Device.objects.filter(
                next_stk__isnull=False,
                next_stk__lte=grenze
            ).count()

        elif widget.widget_type == "mtk":
            widget.value = Device.objects.filter(
                next_mtk__isnull=False,
                next_mtk__lte=grenze
            ).count()

        elif widget.widget_type == "dguv":
            widget.value = Device.objects.filter(
                next_dguv__isnull=False,
                next_dguv__lte=grenze
            ).count()

        elif widget.widget_type == "device_documents":
            widget.value = DeviceDocument.objects.count()

        elif widget.widget_type == "technician_documents":
            widget.value = TechnicianDocument.objects.count()

        else:
            widget.value = 0

    context = {

        "devices": devices,
        "search": search,
        "status": status,

        "offene_reparaturen": offene_reparaturen,
        "reparaturen_bearbeitung": reparaturen_bearbeitung,

        "faellige_stk": faellige_stk,
        "faellige_mtk": faellige_mtk,
        "faellige_dguv": faellige_dguv,

        "luebeck_total": luebeck_total,
        "ratzeburg_total": ratzeburg_total,

        "geraete_uebersicht": geraete_uebersicht,

        "dashboard_settings": dashboard_settings,

        "dashboard_widgets": dashboard_widgets,
    }

    return render(
        request,
        "devices/dashboard.html",
        context
    )


@login_required
def dashboard_settings(request):

    if not request.user.is_superuser:
        return redirect("permission_denied")


    settings, created = DashboardSettings.objects.get_or_create(
        id=1
    )


    if request.method == "POST":

        settings.show_luebeck = request.POST.get(
            "show_luebeck"
        ) == "on"

        settings.show_ratzeburg = request.POST.get(
            "show_ratzeburg"
        ) == "on"

        settings.show_reparaturen = request.POST.get(
            "show_reparaturen"
        ) == "on"

        settings.show_stk = request.POST.get(
            "show_stk"
        ) == "on"

        settings.show_mtk = request.POST.get(
            "show_mtk"
        ) == "on"

        settings.show_dguv = request.POST.get(
            "show_dguv"
        ) == "on"

        settings.show_geraete_uebersicht = request.POST.get(
            "show_geraete_uebersicht"
        ) == "on"


        settings.save()

        messages.success(
            request,
            "Dashboard Einstellungen wurden erfolgreich gespeichert."
        )

        return redirect("dashboard_settings")


    return render(
        request,
        "devices/dashboard_settings.html",
        {
            "settings": settings
        }
    )


@login_required
def upload_document(request, device_id):

    if not request.user.is_superuser:
        return redirect("permission_denied")


    device = get_object_or_404(
        Device,
        id=device_id
    )
    if request.method == "POST":

        form = DeviceDocumentForm(
            request.POST,
            request.FILES
        )

        if form.is_valid():

            document = form.save(commit=False)

            document.device = device

            document.save()

            return redirect(
                "device_detail",
                device_id=device.id
            )

    else:

        form = DeviceDocumentForm()

    return render(
        request,
        "devices/upload_document.html",
        {
            "device": device,
            "form": form,
        }
    )

@login_required
def device_documents_list(request, document_type):

    documents = DeviceDocument.objects.filter(
        document_type=document_type
    ).select_related(
        "device"
    ).order_by(
        "-upload_date"
    )


    return render(
        request,
        "devices/device_documents_list.html",
        {
            "documents": documents,
            "document_type": document_type
        }
    )
@login_required
def device_delete(request, device_id):

    if not request.user.is_superuser:
        return redirect("permission_denied")


    device = get_object_or_404(
        Device,
        id=device_id
    )


    if request.method == "POST":


        AuditLog.objects.create(

            user=request.user,

            action="DELETE",

            model_name="Gerät",

            object_id=device.id,

            description=(
                f"Gerät {device.name} "
                f"(Inventarnummer: {device.inventory_number}, "
                f"Seriennummer: {device.serial_number}) "
                "gelöscht."
            )

        )


        device.delete()


        messages.success(
            request,
            "Gerät erfolgreich gelöscht."
        )


        return redirect(
            "device_list"
        )


    return render(

        request,

        "devices/device_delete_confirm.html",

        {
            "device": device
        }

    )
@login_required
def device_edit(request, device_id):

    if not request.user.is_superuser:
        return redirect("permission_denied")


    device = get_object_or_404(
        Device,
        id=device_id
    )


    if request.method == "POST":


        form = DeviceForm(
            request.POST,
            request.FILES,
            instance=device,
        )


        if form.is_valid():


            device = form.save()


            AuditLog.objects.create(

                user=request.user,

                action="UPDATE",

                model_name="Gerät",

                object_id=device.id,

                description=(
                    f"Gerät {device.name} "
                    f"(Inventarnummer: {device.inventory_number}, "
                    f"Seriennummer: {device.serial_number}) "
                    "geändert."
                )

            )


            messages.success(
                request,
                "Gerät erfolgreich geändert."
            )


            return redirect(
                "device_detail",
                device_id=device.id
            )


    else:


        form = DeviceForm(
            instance=device,
        )


    return render(

        request,

        "devices/device_form.html",

        {
            "form": form,
            "device": device,
        }

    )
@login_required
def document_delete(request, document_id):

    if not request.user.is_superuser:
        return redirect("permission_denied")


    document = get_object_or_404(
        DeviceDocument,
        id=document_id
    )


    document_type = document.document_type


    if request.method == "POST":

        if document.file:
            document.file.delete()

        document.delete()


        return redirect(
            "documents_by_type",
            document_type_id=document_type.id
        )


    return render(
        request,
        "devices/document_delete_confirm.html",
        {
            "document": document
        }
    )
@login_required
def document_rename(request, document_id):

    if not request.user.is_superuser:
        return redirect("permission_denied")


    document = get_object_or_404(
        DeviceDocument,
        id=document_id
    )

    if request.method == "POST":

        new_name = request.POST.get("new_name")

        if new_name:

            document.display_name = new_name

            document.save()

        return redirect(
            "device_detail",
            device_id=document.device.id
        )


    return render(
        request,
        "devices/document_rename.html",
        {
            "document": document
        }
    )

def documents(request):

    return render(
        request,
        "devices/documents.html"
    )

@login_required
def documents_by_type(request, document_type_id):

    document_type = get_object_or_404(
        DocumentType,
        id=document_type_id
    )

    documents = DeviceDocument.objects.filter(
        document_type=document_type
    ).select_related(
        "device",
        "device__geraetart",
        "document_type"
    )

    geraetart = request.GET.get("geraetart")

    if geraetart:
        documents = documents.filter(
            device__geraetart_id=geraetart
        )

    geraetarten = Geraetart.objects.filter(
        aktiv=True
    ).order_by("name")

    standorte = Standort.objects.all()

    return render(
        request,
        "devices/documents_by_type.html",
        {
            "documents": documents.order_by("-upload_date"),
            "document_type": document_type,
            "standorte": standorte,
            "geraetarten": geraetarten,
            "geraetart": geraetart,
        }
    )
@login_required
def technician_documents(request):

    documents = TechnicianDocument.objects.all().order_by("-upload_date")

    return render(
        request,
        "devices/technician_documents.html",
        {
            "documents": documents
        }
    )

@login_required
def technician_document_upload(request):

    if not request.user.is_superuser:
        return redirect("permission_denied")

    if request.method == "POST":

        form = TechnicianDocumentForm(
            request.POST,
            request.FILES
        )

        if form.is_valid():

            document = form.save(commit=False)
            document.uploaded_by = request.user
            document.save()

            return redirect("technician_documents")

    else:

        form = TechnicianDocumentForm()

    return render(
        request,
        "devices/technician_document_upload.html",
        {
            "form": form
        }
    )

@login_required
def technician_document_rename(request, doc_id):

    if not request.user.is_superuser:
        return redirect("permission_denied")

    doc = get_object_or_404(
        TechnicianDocument,
        id=doc_id
    )

    if request.method == "POST":

        new_name = request.POST.get("new_name")

        if new_name:
            doc.title = new_name
            doc.save()

        return redirect("technician_documents")

    return render(
        request,
        "devices/technician_document_rename.html",
        {
            "doc": doc
        }
    )


@login_required
def technician_document_delete(request, doc_id):

    if not request.user.is_superuser:
        return redirect("permission_denied")

    doc = get_object_or_404(
        TechnicianDocument,
        id=doc_id
    )

    if request.method == "POST":

        if doc.file:
            doc.file.delete(save=False)

        doc.delete()

        return redirect("technician_documents")

    return render(
        request,
        "devices/technician_document_delete.html",
        {
            "doc": doc
        }
    )

@login_required
def general_documents(request, category):

    documents = GeneralDocument.objects.filter(
        geraetart=geraetart
    ).order_by("-upload_date")


    return render(
        request,
        "devices/general_documents.html",
        {
            "documents": documents,
             "category": category
        }
    )


@login_required
def general_document_upload(request):
    
    category = request.GET.get("category", "Zertifikat")

    if request.method == "POST":

        form = GeneralDocumentForm(request.POST, request.FILES)

        if form.is_valid():

            document = form.save(commit=False)
            document.category = category
            document.save()

            return redirect(
                "general_documents",
                category=document.category
            )

    else:

        form = GeneralDocumentForm(
            initial={
                "category": category
            }
        )

    return render(
        request,
        "devices/general_upload.html",
        {
            "form": form,
            "category": category,
        }
    )
@login_required
def general_document_rename(request, doc_id):

    if not request.user.is_superuser:
        return redirect("permission_denied")
    doc = get_object_or_404(
        GeneralDocument,
        id=doc_id
    )

    if request.method == "POST":

        new_name = request.POST.get("new_name")

        if new_name:
            doc.display_name = new_name
            doc.save()

        return redirect(
            "general_documents",
            category=doc.category
        )

    return render(
        request,
        "devices/general_document_rename.html",
        {
            "doc": doc
        }
    )

@login_required
def general_document_delete(request, doc_id):

    doc = get_object_or_404(
        GeneralDocument,
        id=doc_id
    )


    if request.method == "POST":

        doc.file.delete()
        doc.delete()

        return redirect(
            "general_documents",
            category=doc.category
        )


    return render(
        request,
        "devices/general_document_delete.html",
        {
            "doc": doc
        }
    )


@login_required
def system_settings(request):


    if request.method == "POST":

        theme = request.POST.get("theme")

        request.session["theme"] = theme

        messages.success(
            request,
            "Systemeinstellungen gespeichert"
        )

        return redirect("settings")


    return render(
        request,
        "devices/system_settings.html"
    )


@login_required
def contact_image_change(request):

    if not request.user.is_superuser:
        return redirect("permission_denied")


    settings = PracticeSettings.objects.first()


    if request.method == "POST":

        if request.FILES.get("contact_image"):

            settings.contact_image = request.FILES["contact_image"]
            settings.save()


        return redirect("contact")


    return render(
        request,
        "devices/contact_image.html",
        {
            "settings": settings
        }
    )

@login_required
def contact_edit(request):

    if not request.user.is_superuser:
        return redirect("permission_denied")


    settings, created = PracticeSettings.objects.get_or_create(
        id=1,
        defaults={
            "practice_name": "Medizintechnik",
            "contact_person": "Praxis",
            "email": "",
            "phone": "",
            "address": "",
        }
    )


    if request.method == "POST":

        settings.contact_name = request.POST.get(
            "name",
            ""
        )

        settings.email = request.POST.get(
            "email",
            ""
        )

        settings.phone = request.POST.get(
            "phone",
            ""
        )

        settings.address = request.POST.get(
            "address",
            ""
        )


        if request.FILES.get("contact_image"):

            settings.contact_image = request.FILES[
                "contact_image"
            ]


        settings.save()


        messages.success(
            request,
            "Kontaktdaten erfolgreich geändert."
        )


        return redirect("contact")


    return render(
        request,
        "devices/contact_edit.html",
        {
            "settings": settings,
        }
    )



@login_required
def user_list(request):

    if not request.user.is_superuser:
        return redirect("permission_denied")

    users = User.objects.all().order_by("username")

    return render(
        request,
        "devices/user_list.html",
        {
            "users": users
        }
    )


@login_required
def user_create(request):

    if not request.user.is_superuser:
        return redirect("home")

    if request.method == "POST":

        username = request.POST.get("username")
        first_name = request.POST.get("first_name")
        last_name = request.POST.get("last_name")
        email = request.POST.get("email")
        password1 = request.POST.get("password1")
        password2 = request.POST.get("password2")
        role = request.POST.get("role")

        if password1 != password2:
            messages.error(request, "Passwörter stimmen nicht überein.")
            return render(request, "devices/user_create.html")

        if User.objects.filter(username=username).exists():
            messages.error(request, "Benutzername existiert bereits.")
            return render(request, "devices/user_create.html")

        user = User.objects.create_user(
            username=username,
            password=password1,
            first_name=first_name,
            last_name=last_name,
            email=email
        )

        if role == "admin":
            user.is_staff = True
            user.is_superuser = True
        else:
            user.is_staff = False
            user.is_superuser = False

        user.save()

        messages.success(
            request,
            "Benutzer erfolgreich erstellt."
        )

        return redirect("user_list")

    return render(
        request,
        "devices/user_create.html"
    )

@login_required
def user_edit(request, user_id):

    if not request.user.is_superuser:
        return redirect("home")

    user = get_object_or_404(User, id=user_id)

    if request.method == "POST":

        user.username = request.POST.get("username")
        user.first_name = request.POST.get("first_name")
        user.last_name = request.POST.get("last_name")
        user.email = request.POST.get("email")

        role = request.POST.get("role")

        if role == "admin":
            user.is_staff = True
            user.is_superuser = True
        else:
            user.is_staff = False
            user.is_superuser = False

        user.save()

        messages.success(request, "Benutzer erfolgreich geändert.")

        return redirect("user_list")

    return render(
        request,
        "devices/user_edit.html",
        {
            "edit_user": user
        }
    )




@login_required
def user_delete(request, user_id):

    if not request.user.is_superuser:
        return redirect("home")

    user = get_object_or_404(User, id=user_id)

    # منع حذف المستخدم الحالي
    if user == request.user:
        messages.error(
            request,
            "Sie können Ihren eigenen Benutzer nicht löschen."
        )
        return redirect("user_list")

    # منع حذف آخر Administrator
    if user.is_superuser:
        admins = User.objects.filter(is_superuser=True).count()

        if admins <= 1:
            messages.error(
                request,
                "Der letzte Administrator kann nicht gelöscht werden."
            )
            return redirect("user_list")

    user.delete()

    messages.success(
        request,
        "Benutzer erfolgreich gelöscht."
    )

    return redirect("user_list")

@login_required
def user_password_reset(request, user_id):

    if not request.user.is_superuser:
        return redirect("home")

    user = get_object_or_404(User, id=user_id)

    if request.method == "POST":

        password1 = request.POST.get("password1")
        password2 = request.POST.get("password2")

        if password1 != password2:
            messages.error(
                request,
                "Passwörter stimmen nicht überein."
            )
            return redirect(
                "user_password_reset",
                user_id=user.id
            )

        user.set_password(password1)
        user.save()

        messages.success(
            request,
            "Passwort erfolgreich geändert."
        )

        return redirect("user_list")


    return render(
        request,
        "registration/user_password_reset.html",
        {
            "edit_user": user
        }
    )

@login_required
def permission_denied(request):
    return render(
        request,
        "devices/permission_denied.html"
    )


@login_required
def profile_settings(request):

    return render(
        request,
        "devices/profile_settings.html"
    )


@login_required
def system_management(request):

    return render(
        request,
        "devices/system_management.html"
    )

@login_required
def standort_list(request):

    standorte = Standort.objects.all().order_by("name")

    return render(
        request,
        "devices/standort_list.html",
        {
            "standorte": standorte
        }
    )


@login_required
def standort_create(request):

    if not request.user.is_superuser:
        return redirect("permission_denied")

    if request.method == "POST":

        Standort.objects.create(
            name=request.POST.get("name"),
            address=request.POST.get("address"),
            phone=request.POST.get("phone"),
            email=request.POST.get("email"),
            active=True if request.POST.get("active") else False,
        )

        return redirect("standort_list")

    return render(
        request,
        "devices/standort_create.html"
    )

@login_required
def standort_edit(request, standort_id):

    if not request.user.is_superuser:
        return redirect("permission_denied")

    standort = get_object_or_404(
        Standort,
        id=standort_id
    )

    if request.method == "POST":

        standort.name = request.POST.get("name")
        standort.address = request.POST.get("address")
        standort.phone = request.POST.get("phone")
        standort.email = request.POST.get("email")
        standort.active = True if request.POST.get("active") else False

        standort.save()

        return redirect("standort_list")

    return render(
        request,
        "devices/standort_edit.html",
        {
            "standort": standort
        }
    )

@login_required
def standort_delete(request, standort_id):

    if not request.user.is_superuser:
        return redirect("permission_denied")

    standort = get_object_or_404(
        Standort,
        id=standort_id
    )

    # لا يمكن حذف الفرع إذا كان مستخدمًا
    if Device.objects.filter(practice=standort).exists():

        messages.error(
            request,
            "Dieser Standort wird von Geräten verwendet und kann nicht gelöscht werden."
        )

        return redirect("standort_list")

    standort.delete()

    return redirect("standort_list")

@login_required
def login_view(request):

    if request.method == "POST":

        username = request.POST.get("username")
        password = request.POST.get("password")

        user = authenticate(
            request,
            username=username,
            password=password
        )

        if user is not None:
            login(request, user)
            return redirect("home")

        else:
            return render(
                request,
                "registration/login.html",
                {"error": "Benutzername oder Passwort falsch"}
            )

    return render(
        request,
        "registration/login.html"
    )



@login_required
def documents_home(request):

    document_types = DocumentType.objects.all().order_by("order")

    return render(
        request,
        "devices/documents_home.html",
        {
            "document_types": document_types
        }
    )


@login_required
def filterwechsel_create(request):

    if request.method == "POST":

        form = FilterwechselForm(request.POST)


        if form.is_valid():

            filterwechsel = form.save()


            AuditLog.objects.create(

                user=request.user,

                action="CREATE",

                model_name="Filterwechsel",

                object_id=filterwechsel.id,

                description=(

                    f"Filterwechsel erstellt. "

                    f"Gerät: "
                    f"{filterwechsel.geraet.name if filterwechsel.geraet else 'Unbekannt'} "

                    f"(Inventarnummer: "
                    f"{filterwechsel.geraet.inventory_number if filterwechsel.geraet else '—'})"

                )

            )


            messages.success(
                request,
                "Filterwechsel erfolgreich gespeichert."
            )


            return redirect(
                "filterwechsel_history"
            )


    else:

        form = FilterwechselForm()


    return render(
        request,
        "devices/filterwechsel_create.html",
        {
            "form": form
        }
    )
@login_required
def filterwechsel_history(request):

    filterwechsel = Filterwechsel.objects.all().order_by("-datum")

    search = request.GET.get("search")

    if search:
        filterwechsel = filterwechsel.filter(
            inventarnummer__icontains=search
        )

    return render(
        request,
        "devices/filterwechsel_history.html",
        {
            "filterwechsel": filterwechsel,
            "search": search or ""
        }
    )





@login_required
def filterwechsel_update(request, id):

    filterwechsel = get_object_or_404(
        Filterwechsel,
        id=id
    )


    original_datum = filterwechsel.datum



    if request.method == "POST":


        form = FilterwechselForm(

            request.POST,

            instance=filterwechsel

        )


        if form.is_valid():


            obj = form.save(
                commit=False
            )


            obj.datum = original_datum


            obj.save()



            AuditLog.objects.create(

                user=request.user,

                action="UPDATE",

                model_name="Filterwechsel",

                object_id=obj.id,

                description=(

                    f"Filterwechsel geändert. "

                    f"Gerät: "
                    f"{obj.geraet.name if obj.geraet else 'Unbekannt'} "

                    f"(Inventarnummer: "
                    f"{obj.geraet.inventory_number if obj.geraet else '—'})"

                )

            )



            messages.success(

                request,

                "Filterwechsel geändert."

            )


            return redirect(
                "filterwechsel_history"
            )


    else:


        form = FilterwechselForm(
            instance=filterwechsel
        )



    return render(

        request,

        "devices/filterwechsel_create.html",

        {
            "form": form,
            "edit": True
        }

    )

@login_required
def filterwechsel_delete(request, id):
    if not request.user.is_superuser:
        return redirect("permission_denied")
    filterwechsel = get_object_or_404(
        Filterwechsel,
        id=id
    )


    if request.method == "POST":


        AuditLog.objects.create(

            user=request.user,

            action="DELETE",

            model_name="Filterwechsel",

            object_id=filterwechsel.id,

            description=(

                f"Filterwechsel gelöscht. "

                f"Gerät: "
                f"{filterwechsel.geraet.name if filterwechsel.geraet else 'Unbekannt'} "

                f"(Inventarnummer: "
                f"{filterwechsel.geraet.inventory_number if filterwechsel.geraet else '—'})"

            )

        )


        filterwechsel.delete()


        messages.success(
            request,
            "Filterwechsel gelöscht."
        )


    return redirect(
        "filterwechsel_history"
    )

@login_required
def geraetarten(request):

    geraetarten = Geraetart.objects.all().order_by("name")

    return render(
        request,
        "devices/geraetarten.html",
        {
            "geraetarten": geraetarten,
        },
    )

@login_required
def geraetart_create(request):

    if not request.user.is_superuser:
        return redirect("permission_denied")

    if request.method == "POST":

        form = GeraetartForm(request.POST)

        if form.is_valid():
            form.save()
            return redirect("geraetarten")

    else:
        form = GeraetartForm()

    return render(
        request,
        "devices/geraetart_form.html",
        {
            "form": form,
            "title": "Gerätart hinzufügen",
        }
    )


@login_required
def geraetart_update(request, id):

    if not request.user.is_superuser:
        return redirect("permission_denied")

    geraetart = get_object_or_404(
        Geraetart,
        id=id
    )

    if request.method == "POST":

        form = GeraetartForm(
            request.POST,
            instance=geraetart
        )

        if form.is_valid():
            form.save()
            return redirect("geraetarten")

    else:

        form = GeraetartForm(
            instance=geraetart
        )

    return render(
        request,
        "devices/geraetart_form.html",
        {
            "form": form,
            "title": "Gerätart bearbeiten",
        }
    )

@login_required
def geraetart_delete(request, id):

    if not request.user.is_superuser:
        return redirect("permission_denied")

    geraetart = get_object_or_404(
        Geraetart,
        id=id
    )

    geraetart.delete()

    return redirect("geraetarten")

@login_required
def document_types(request):
    
    types = DocumentType.objects.all()

    return render(
        request,
        "devices/document_types.html",
        {
            "types": types
        }
    )

@login_required
def add_document_type(request):
    if not request.user.is_superuser:
        return redirect("permission_denied")

    if request.method == "POST":
        name = request.POST.get("name")

        if name:
            last_order = DocumentType.objects.count()

            DocumentType.objects.create(
                name=name,
                order=last_order
            )

            return redirect("document_types")

    return render(
        request,
        "devices/add_document_type.html"
    )

@login_required
def edit_document_type(request, id):
    if not request.user.is_superuser:
            return redirect("permission_denied")

    document_type = DocumentType.objects.get(id=id)

    if request.method == "POST":

        document_type.name = request.POST.get("name")
        document_type.save()

        return redirect("document_types")

    return render(
        request,
        "devices/edit_document_type.html",
        {
            "type": document_type
        }
    )
@login_required
def delete_document_type(request, id):

    document_type = DocumentType.objects.get(id=id)

    document_type.delete()

    return redirect("document_types")


@login_required
def backup(request):

    if not request.user.is_superuser:
        return redirect("permission_denied")

    backup_dir = os.path.join(settings.BASE_DIR, "backups")
    os.makedirs(backup_dir, exist_ok=True)

    if request.method == "POST":

        db_path = os.path.join(settings.BASE_DIR, "db.sqlite3")

        filename = datetime.now().strftime(
            "backup_%Y-%m-%d_%H-%M-%S.sqlite3"
        )

        destination = os.path.join(
            backup_dir,
            filename
        )

        shutil.copy2(
            db_path,
            destination
        )

        messages.success(
            request,
            "Backup erfolgreich erstellt."
        )

        return redirect("backup")

    backups = []

    for file in os.listdir(backup_dir):

        if file.endswith(".sqlite3"):

            path = os.path.join(backup_dir, file)

            backups.append({
                "name": file,
                "size": round(os.path.getsize(path) / 1024 / 1024, 2),
                "date": datetime.fromtimestamp(
                    os.path.getmtime(path)
                ),
            })

    backups.sort(
        key=lambda x: x["date"],
        reverse=True
    )

    return render(
        request,
        "devices/backup.html",
        {
            "backups": backups,
        }
    )

@login_required
def backup_download(request, filename):

    if not request.user.is_superuser:
        return redirect("permission_denied")

    path = os.path.join(
        settings.BASE_DIR,
        "backups",
        filename
    )

    if not os.path.exists(path):
        raise Http404("Backup nicht gefunden.")

    return FileResponse(
        open(path, "rb"),
        as_attachment=True,
        filename=filename
    )


@login_required
def backup_delete(request, filename):

    if not request.user.is_superuser:
        return redirect("permission_denied")

    path = os.path.join(
        settings.BASE_DIR,
        "backups",
        filename
    )

    if os.path.exists(path):
        os.remove(path)
        messages.success(
            request,
            "Backup erfolgreich gelöscht."
        )
    else:
        messages.error(
            request,
            "Backup nicht gefunden."
        )

    return redirect("backup")

@login_required
def backup_restore(request, filename):

    if not request.user.is_superuser:
        return redirect("permission_denied")

    if request.method != "POST":
        return redirect("backup")

    backup_path = os.path.join(
        settings.BASE_DIR,
        "backups",
        filename
    )

    db_path = os.path.join(
        settings.BASE_DIR,
        "db.sqlite3"
    )

    if not os.path.exists(backup_path):
        messages.error(
            request,
            "Backup nicht gefunden."
        )
        return redirect("backup")

    shutil.copy2(
        backup_path,
        db_path
    )

    messages.success(
        request,
        "Backup erfolgreich wiederhergestellt. Bitte starten Sie den Server neu."
    )

    return redirect("backup")


def create_daily_backup(request=None):

    backup_dir = os.path.join(settings.BASE_DIR, "backups")
    os.makedirs(backup_dir, exist_ok=True)

    today = datetime.now().strftime("%Y-%m-%d")

    filename = f"daily_{today}.sqlite3"

    destination = os.path.join(
        backup_dir,
        filename
    )

    if not os.path.exists(destination):

        db = os.path.join(
            settings.BASE_DIR,
            "db.sqlite3"
        )

        shutil.copy2(
            db,
            destination
        )
        return filename


def create_update_backup():

    backup_dir = os.path.join(
        settings.BASE_DIR,
        "backups"
    )

    os.makedirs(
        backup_dir,
        exist_ok=True
    )

    db_path = os.path.join(
        settings.BASE_DIR,
        "db.sqlite3"
    )

    filename = datetime.now().strftime(
        "update_backup_%Y-%m-%d_%H-%M-%S.sqlite3"
    )

    destination = os.path.join(
        backup_dir,
        filename
    )

    shutil.copy2(
        db_path,
        destination
    )

    return filename


@login_required
def export(request):

    if not request.user.is_superuser:
        return redirect("permission_denied")
    company = CompanyInformation.objects.first()

    export_dir = os.path.join(
        settings.BASE_DIR,
        "exports"
    )

    os.makedirs(
        export_dir,
        exist_ok=True
    )


    if request.method == "POST":

        
        geraetart_filter = request.POST.get(
            "geraetart"
        )

        standort_filter = request.POST.get(
            "standort"
        )


        devices = Device.objects.all()


        if geraetart_filter:

            devices = devices.filter(
                geraetart__name=geraetart_filter
            )


        if standort_filter:

            devices = devices.filter(
                practice__name=standort_filter
            )



        wb = Workbook()

        ws = wb.active

        ws.title = "Geräte"

        # Firmeninformationen

        if company:

            ws.append([
                company.company_name
            ])

            ws.append([
                company.address
            ])

            ws.append([
                f"Telefon: {company.phone}"
            ])

            ws.append([
                f"E-Mail: {company.email}"
            ])

            ws.append([
                f"Kundennummer: {company.customer_number}"
            ])


        ws.append([])

        # Bericht Kopf

        ws.append([
            "Geräte Export"
        ])


        ws.append([
            f"Erstellt am: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
        ])


        ws.append([
            f"Geräteart: {geraetart_filter if geraetart_filter else 'Alle Geräte'}"
        ])


        ws.append([
            f"Standort: {standort_filter if standort_filter else 'Alle Standorte'}"
        ])


        ws.append([])



        # Tabellen Kopf

        ws.append([

            "Inventarnummer",

            "Seriennummer",

            "Gerätname",

            "Gerätart",

            "Standort",

            "Betriebsstunden",

            "Letzte STK",

            "Letzte MTK",

            "Letzte DGUV V3",

        ])



        # Kopf formatieren

        for cell in ws[6]:

            cell.font = Font(
                bold=True
            )

            cell.alignment = Alignment(
                horizontal="center"
            )



        for device in devices:


            betriebsstunden = ""


            if (
                device.geraetart
                and device.geraetart.name == "Dialyse Maschinen"
            ):

                betriebsstunden = (
                    device.operating_hours
                    or ""
                )



            ws.append([


                device.inventory_number,


                device.serial_number,


                device.name,


                device.geraetart.name
                if device.geraetart
                else "",


                device.practice.name
                if device.practice
                else "",


                betriebsstunden,


                device.last_stk.strftime("%d.%m.%Y")
                if device.last_stk
                else "",


                device.last_mtk.strftime("%d.%m.%Y")
                if device.last_mtk
                else "",


                device.last_dguv.strftime("%d.%m.%Y")
                if device.last_dguv
                else "",


            ])




        # Spalten automatisch anpassen

        for column in ws.columns:

            max_length = 0

            column_letter = get_column_letter(
                column[0].column
            )


            for cell in column:

                if cell.value:

                    max_length = max(
                        max_length,
                        len(str(cell.value))
                    )


            ws.column_dimensions[column_letter].width = (
                max_length + 3
            )



        filename = datetime.now().strftime(
            "Geraete_Export_%Y-%m-%d_%H-%M-%S.xlsx"
        )


        filepath = os.path.join(
            export_dir,
            filename
        )


        wb.save(filepath)



        messages.success(
            request,
            "Export erfolgreich erstellt."
        )


        return redirect("export")




    exports = []


    for file in os.listdir(export_dir):

        path = os.path.join(
            export_dir,
            file
        )


        exports.append({

            "name": file,

            "date": datetime.fromtimestamp(
                os.path.getmtime(path)
            ),

            "size": round(
                os.path.getsize(path) / 1024 / 1024,
                2
            ),

        })



    exports.sort(
        key=lambda x: x["date"],
        reverse=True
    )



    return render(
        request,
        "devices/export.html",
        {

            "exports": exports,

            "geraetearten": Geraetart.objects.all(),

            "standorte": Standort.objects.all(),

        }
    )

@login_required
def export_download(request, filename):

    if not request.user.is_superuser:
        return redirect("permission_denied")


    export_dir = os.path.join(
        settings.BASE_DIR,
        "exports"
    )


    filepath = os.path.join(
        export_dir,
        filename
    )


    if not os.path.exists(filepath):

        messages.error(
            request,
            "Export-Datei nicht gefunden."
        )

        return redirect("export")


    return FileResponse(
        open(filepath, "rb"),
        as_attachment=True,
        filename=filename
    )

@login_required
def export_delete(request, filename):

    if not request.user.is_superuser:
        return redirect("permission_denied")


    export_dir = os.path.join(
        settings.BASE_DIR,
        "exports"
    )


    filepath = os.path.join(
        export_dir,
        filename
    )


    if os.path.exists(filepath):

        os.remove(filepath)

        messages.success(
            request,
            "Export erfolgreich gelöscht."
        )

    else:

        messages.error(
            request,
            "Export-Datei nicht gefunden."
        )


    return redirect("export")

@login_required
def firmeninformationen_list(request):

    search = request.GET.get(
        "search",
        ""
    )


    companies = CompanyInformation.objects.all()


    if search:

        companies = companies.filter(
            Q(company_name__icontains=search) |
            Q(address__icontains=search)
        )


    return render(
        request,
        "devices/firmeninformationen_list.html",
        {
            "companies": companies,
            "search": search
        }
    )



@login_required
def firmeninformationen_manage(request):

    
    companies = CompanyInformation.objects.all()


    return render(
        request,
        "devices/firmeninformationen_manage.html",
        {
            "companies": companies
        }
    )



@login_required
def firmeninformationen_create(request):

    if not request.user.is_superuser:
        return redirect("permission_denied")


    if request.method == "POST":

        company = CompanyInformation.objects.create(

            company_name=request.POST.get(
                "company_name"
            ),

            address=request.POST.get(
                "address"
            ),

            phone=request.POST.get(
                "phone"
            ),

            email=request.POST.get(
                "email"
            ),

            customer_number=request.POST.get(
                "customer_number"
            )

        )

        AuditLog.objects.create(

            user=request.user,

            action="CREATE",

            model_name="Firma",

            object_id=company.id,

            description=f"Firma {company.company_name} hinzugefügt."

        )
        messages.success(
            request,
            "Firma erfolgreich hinzugefügt."
        )


        return redirect(
            "firmeninformationen_manage"
        )


    return render(
        request,
        "devices/firmeninformation_create.html"
    )



@login_required
def firmeninformationen_edit(request, id):


    company = CompanyInformation.objects.get(
        id=id
    )


    if request.method == "POST":

        company.company_name = request.POST.get(
            "company_name"
        )

        company.address = request.POST.get(
            "address"
        )

        company.phone = request.POST.get(
            "phone"
        )

        company.email = request.POST.get(
            "email"
        )

        company.customer_number = request.POST.get(
            "customer_number"
        )


        company.save()

        AuditLog.objects.create(

            user=request.user,

            action="UPDATE",

            model_name="Firma",

            object_id=company.id,

            description=f"Firma {company.company_name} geändert."

        )
        messages.success(
            request,
            "Firma geändert."
        )


        return redirect(
            "firmeninformationen_manage"
        )


    return render(
        request,
        "devices/firmeninformationen_edit.html",
        {
            "company": company
        }
    )

@login_required
def firmeninformationen_delete(request, id):

    if not request.user.is_superuser:
        return redirect("permission_denied")


    company = CompanyInformation.objects.get(
        id=id
    )


    if request.method == "POST":

        AuditLog.objects.create(

            user=request.user,

            action="DELETE",

            model_name="Firma",

            object_id=company.id,

            description=f"Firma {company.company_name} gelöscht."

        )
        company.delete()


        messages.success(
            request,
            "Firma erfolgreich gelöscht."
        )


    return redirect(
        "firmeninformationen_manage"
    )

@login_required
def audit_log(request):

    if not request.user.is_superuser:
        return redirect("permission_denied")


    logs = AuditLog.objects.all()


    search = request.GET.get(
        "search",
        ""
    )


    if search:

        logs = logs.filter(
            Q(user__username__icontains=search) |
            Q(model_name__icontains=search) |
            Q(description__icontains=search)
        )


    return render(
        request,
        "devices/audit_log.html",
        {
            "logs": logs,
            "search": search
        }
    )

@login_required
def audit_log_delete(request, id):

    if not request.user.is_superuser:
        return redirect("permission_denied")


    log = get_object_or_404(
        AuditLog,
        id=id
    )


    if request.method == "POST":

        log.delete()


        messages.success(
            request,
            "Audit Log Eintrag gelöscht."
        )


    return redirect(
        "audit_log"
    )

@login_required
def audit_log_delete_selected(request):

    if not request.user.is_superuser:
        return redirect("permission_denied")


    if request.method == "POST":

        ids = request.POST.getlist("logs")


        if ids:

            AuditLog.objects.filter(
                id__in=ids
            ).delete()


            messages.success(
                request,
                "Ausgewählte Audit Logs wurden gelöscht."
            )


    return redirect(
        "audit_log"
    )

@login_required
def kontakt_informationen(request):

    if not request.user.is_superuser:
        return redirect("permission_denied")


    kontakt, created = ContactInformation.objects.get_or_create(
        id=1
    )


    if request.method == "POST":

        kontakt.technician_name = request.POST.get(
            "technician_name"
        )

        kontakt.email = request.POST.get(
            "email"
        )

        kontakt.phone = request.POST.get(
            "phone"
        )

        kontakt.address = request.POST.get(
            "address"
        )


        if request.FILES.get("image"):

            kontakt.image = request.FILES.get(
                "image"
            )


        kontakt.save()


        messages.success(
            request,
            "Kontaktdaten gespeichert."
        )


        return redirect(
            "kontakt"
        )


    return render(
        request,
        "devices/kontakt_informationen.html",
        {
            "kontakt": kontakt
        }
    )

@login_required
def kontakt_edit(request):

    if not request.user.is_superuser:
        return redirect("permission_denied")

    kontakt, created = ContactInformation.objects.get_or_create(id=1)

    if request.method == "POST":

        kontakt.technician_name = request.POST.get("technician_name")
        kontakt.email = request.POST.get("email")
        kontakt.phone = request.POST.get("phone")
        kontakt.address = request.POST.get("address")

        kontakt.save()

        # حذف الصور المحددة
        if "delete_selected" in request.POST:

            delete_ids = request.POST.getlist("delete_images")

            ContactImage.objects.filter(
                id__in=delete_ids,
                contact=kontakt
            ).delete()

        # إضافة الصور الجديدة
        images = request.FILES.getlist("images")

        for image in images:

            ContactImage.objects.create(
                contact=kontakt,
                image=image
            )

        messages.success(
            request,
            "Kontaktdaten gespeichert."
        )

        return redirect("kontakt_edit")

    return render(
        request,
        "devices/kontakt_informationen.html",
        {
            "kontakt": kontakt
        }
    )

@login_required
def dashboard_widgets(request):

    if not request.user.is_superuser:
        return redirect("permission_denied")

    widgets = DashboardWidget.objects.all()

    return render(
        request,
        "devices/dashboard_widgets.html",
        {
            "widgets": widgets,
        },
    )

@login_required
def dashboard_widget_create(request):

    if not request.user.is_superuser:
        return redirect("permission_denied")

    if request.method == "POST":

        widget = DashboardWidget()

        widget.widget_type = request.POST.get("widget_type")

        widget.title = request.POST.get("title")

        standort_id = request.POST.get("standort")

        if standort_id:
            widget.standort = Standort.objects.get(id=standort_id)

        geraetart_id = request.POST.get("geraetart")

        if geraetart_id:
            widget.geraetart = Geraetart.objects.get(id=geraetart_id)

        widget.visible = (
            request.POST.get("visible") == "on"
        )

        widget.order = request.POST.get("order") or 0

        widget.color = request.POST.get("color") or "primary"

        widget.icon = request.POST.get("icon") or "bi-grid"

        widget.save()

        messages.success(
            request,
            "Dashboard Widget erstellt."
        )

        return redirect("dashboard_widgets")

    return render(
        request,
        "devices/dashboard_widget_form.html",
        {
            "standorte": Standort.objects.filter(active=True),
            "geraetarten": Geraetart.objects.filter(aktiv=True),
            "types": DashboardWidget.WIDGET_TYPES,
        }
    )

@login_required
def dashboard_widget_edit(request, id):

    if not request.user.is_superuser:
        return redirect("permission_denied")


    widget = get_object_or_404(
        DashboardWidget,
        id=id
    )


    if request.method == "POST":

        widget.widget_type = request.POST.get(
            "widget_type"
        )


        widget.title = request.POST.get(
            "title"
        )


        standort_id = request.POST.get(
            "standort"
        )

        if standort_id:
            widget.standort = get_object_or_404(
                Standort,
                id=standort_id
            )
        else:
            widget.standort = None



        geraetart_id = request.POST.get(
            "geraetart"
        )

        if geraetart_id:
            widget.geraetart = get_object_or_404(
                Geraetart,
                id=geraetart_id
            )
        else:
            widget.geraetart = None



        widget.color = request.POST.get(
            "color"
        ) or "primary"



        widget.icon = request.POST.get(
            "icon"
        ) or "bi-grid"



        widget.order = request.POST.get(
            "order"
        ) or 0



        widget.visible = (
            request.POST.get("visible") == "on"
        )


        widget.save()


        messages.success(
            request,
            "Dashboard Widget geändert."
        )


        return redirect(
            "dashboard_widgets"
        )



    return render(
        request,
        "devices/dashboard_widget_form.html",
        {
            "widget": widget,

            "standorte": Standort.objects.filter(
                active=True
            ),

            "geraetarten": Geraetart.objects.filter(
                aktiv=True
            ),

            "types": DashboardWidget.WIDGET_TYPES,
        }
    )
@login_required
def dashboard_widget_delete(request, id):

    if not request.user.is_superuser:
        return redirect("permission_denied")


    widget = get_object_or_404(
        DashboardWidget,
        id=id
    )


    widget.delete()


    messages.success(
        request,
        "Dashboard Widget gelöscht."
    )


    return redirect(
        "dashboard_widgets"
    )


@login_required
def system_update(request):

    # ملف آخر إصدار متوفر
    update_file = Path(settings.BASE_DIR) / "update" / "version.json"

    latest_version = None
    latest_description = None
    update_available = False


    # قراءة معلومات التحديث
    if update_file.exists():

        with open(update_file, "r", encoding="utf-8") as f:

            data = json.load(f)

            latest_version = data.get("version")
            latest_description = data.get("description")


        try:
            subprocess.run(
                ["git", "fetch", "origin"],
                cwd=settings.BASE_DIR,
                check=True,
                capture_output=True,
                text=True,
            )

            local_commit = subprocess.check_output(
                ["git", "rev-parse", "HEAD"],
                cwd=settings.BASE_DIR,
                text=True,
            ).strip()

            remote_commit = subprocess.check_output(
                ["git", "rev-parse", "origin/main"],
                cwd=settings.BASE_DIR,
                text=True,
            ).strip()

            update_available = (local_commit != remote_commit)

        except Exception:
            update_available = False



    # قراءة سجل التحديثات
    history_file = (
        Path(settings.BASE_DIR)
        / "update"
        / "logs"
        / "update_history.json"
    )

    update_history = []


    if history_file.exists():

        with open(
            history_file,
            "r",
            encoding="utf-8-sig"
        ) as f:

            update_history = json.load(f)



    # قراءة Update Log

    log_file = (
        Path(settings.BASE_DIR)
        / "update"
        / "logs"
        / "update.log"
    )

    update_log = ""


    if log_file.exists():

        with open(
            log_file,
            "r",
            encoding="utf-8"
        ) as f:

            update_log = f.read()



    print("Current Version:", settings.APP_VERSION)
    print("Latest Version:", latest_version)
    print("Update Available:", update_available)



    return render(
        request,
        "devices/system_update.html",
        {
            "current_version": settings.APP_VERSION,
            "latest_version": latest_version,
            "latest_description": latest_description,
            "update_available": update_available,
            "update_history": update_history,
            "update_log": update_log,
        }
    )

@login_required
def check_update(request):

    if not request.user.is_superuser:
        return redirect("permission_denied")

    try:
        # تحديث معلومات المستودع من GitHub
        subprocess.run(
            ["git", "fetch", "origin"],
            cwd=settings.BASE_DIR,
            check=True,
            capture_output=True,
            text=True
        )

        # الحصول على آخر Commit محلي
        local_commit = subprocess.check_output(
            ["git", "rev-parse", "HEAD"],
            cwd=settings.BASE_DIR,
            text=True
        ).strip()

        # الحصول على آخر Commit على GitHub
        remote_commit = subprocess.check_output(
            ["git", "rev-parse", "origin/main"],
            cwd=settings.BASE_DIR,
            text=True
        ).strip()

        if local_commit != remote_commit:

            backup_file = create_update_backup()

            messages.success(
                request,
                f"New update available.\nBackup created:\n{backup_file}"
            )

        else:

            messages.info(
                request,
                "Your system is already up to date."
            )

    except Exception as e:

        messages.error(
            request,
            f"Update check failed: {e}"
        )

    return redirect("system_update")


@login_required
def run_update(request):

    if not request.user.is_superuser:
        return redirect("permission_denied")


    if request.method != "POST":
        return redirect("system_update")


    update_script = os.path.join(
        settings.BASE_DIR,
        "update",
        "scripts",
        "update.bat"
    )


    if not os.path.exists(update_script):

        messages.error(
            request,
            "Update script not found."
        )

        return redirect("system_update")



    # كتابة Log قبل بدء التحديث

    log_dir = os.path.join(
        settings.BASE_DIR,
        "update",
        "logs"
    )

    os.makedirs(
        log_dir,
        exist_ok=True
    )


    log_file = os.path.join(
        log_dir,
        "update.log"
    )


    with open(
        log_file,
        "a",
        encoding="utf-8"
    ) as f:

        f.write(
            f"\n[{datetime.now()}] Update started from System Update page\n"
        )



    # تشغيل ملف التحديث

    subprocess.Popen(
        [
            "cmd",
            "/c",
            update_script
        ],
        cwd=settings.BASE_DIR
    )


    messages.success(
        request,
        "Update process started successfully."
    )


    return redirect("system_update")